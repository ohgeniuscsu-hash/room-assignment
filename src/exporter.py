import pandas as pd
from openpyxl.styles import PatternFill


def export_result(
    result: pd.DataFrame,
    warnings: list[dict],
    combined_issues: list[dict],
    output_path: str,
) -> None:
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        result.to_excel(writer, sheet_name='배정결과', index=False)
        ws = writer.sheets['배정결과']

        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        for row_idx, val in enumerate(result['배정강의실'], start=2):
            if '검토필요' in str(val):
                for col_idx in range(1, len(result.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

        all_warnings = warnings + combined_issues
        if all_warnings:
            warn_df = pd.DataFrame(all_warnings)
            warn_df.to_excel(writer, sheet_name='경고', index=False)


def export_timetable(result: pd.DataFrame, rooms: pd.DataFrame, output_path: str) -> None:
    from src.parser import parse_period

    days_order = ['월', '화', '수', '목', '금', '토']
    max_period = 9

    slots = [(d, p) for d in days_order for p in range(1, max_period + 1)]
    slot_labels = [f"{d}{p}" for d, p in slots]

    assigned_rooms = result[
        ~result['배정강의실'].isin(['온라인', '시간미배정-검토필요', '미배정-검토필요', ''])
    ]['배정강의실'].unique()

    timetable = pd.DataFrame(index=slot_labels, columns=sorted(assigned_rooms))
    timetable = timetable.fillna('')

    for _, row in result.iterrows():
        room_code = str(row.get('배정강의실', '')).strip()
        if room_code not in timetable.columns:
            continue
        day = str(row.get('요일', '')).strip()
        period_str = str(row.get('교시', '')).strip()
        if not day or not period_str:
            continue
        try:
            for p in parse_period(period_str):
                slot = f"{day}{p}"
                if slot in timetable.index:
                    cell_val = f"{row['과목명']} ({row['학과']}/{row['전공']}, {row['교수명']})"
                    timetable.at[slot, room_code] = cell_val
        except ValueError:
            pass

    timetable.to_excel(output_path)
